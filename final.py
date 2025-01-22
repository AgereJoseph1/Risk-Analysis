import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
from io import BytesIO
from openpyxl.styles import PatternFill

###########################
# 1. CONFIG & THRESHOLDS
###########################
RISK_COLORS = {
    1: ('#006400', 'Very Low'),  # Deep green
    2: ('#90EE90', 'Low'),       # Light green
    3: ('#FFD700', 'Medium'),    # Yellow
    4: ('#FF0000', 'High'),      # Red
    5: ('#722F37', 'Extreme')    # Wine
}

THRESHOLD_RULES = {
    'time': {
        2: 1,    # ≤ 2 days => 1
        3: 2,
        4: 3,
        5: 4,
        float('inf'): 5
    },
    'percentage': {
        10: 1,   # ≤ 10% => 1
        11: 2,
        25: 3,
        26: 4,
        float('inf'): 5
    }
}

st.set_page_config(layout="wide", page_title="Risk Intelligence Dashboard Hub", page_icon="🎯")

# Custom CSS
st.markdown("""
    <style>
    .main{background-color:#f8f9fa;}
    .dashboard-header{padding:2.5rem 2rem;background:linear-gradient(135deg,#1E3F66 0%,#2E5090 50%,#3A63B8 100%);color:white;border-radius:15px;margin:1rem 0 3rem 0;box-shadow:0 8px 20px rgba(0,0,0,0.15);text-align:center;}
    .dashboard-header h1{font-size:2.5rem;font-weight:600;margin-bottom:0.5rem;text-shadow:2px 2px 4px rgba(0,0,0,0.1);}
    .dashboard-subtitle{font-size:1.1rem;opacity:0.9;max-width:600px;margin:0 auto;}
    .metric-card{background:white;padding:1.8rem;border-radius:12px;box-shadow:0 4px 12px rgba(0,0,0,0.08);margin-bottom:1.5rem;border:1px solid rgba(0,0,0,0.05);}
    .stButton>button{background-color:#1E3F66;color:white;border:none;padding:0.75rem 1.5rem;border-radius:5px;font-size:1rem;transition:background-color 0.3s ease;}
    #MainMenu,footer,header{visibility:hidden;}
    .main .block-container{padding-top:2rem;padding-bottom:2rem;}
    </style>
""", unsafe_allow_html=True)

###########################
# 2. DATA PROCESSING
###########################
def process_data(df):
    df = df.copy()

    def determine_limitation(appetite_score):
        if 0 <= appetite_score < 10: return 1
        elif 10 <= appetite_score < 11: return 2
        elif 11 <= appetite_score < 25: return 3
        elif 25 <= appetite_score < 26: return 4
        elif 26 <= appetite_score <= 45: return 5
        else: return 5

    df['Limitation'] = df['Appetite Score'].apply(determine_limitation)
    df['Rating'] = df['Limitation'].map({
        1: RISK_COLORS[1][1],  # Very Low
        2: RISK_COLORS[2][1],  # Low
        3: RISK_COLORS[3][1],  # Medium
        4: RISK_COLORS[4][1],  # High
        5: RISK_COLORS[5][1]   # Extreme
    })
    df['Risk_Score'] = df['Limitation'] * df['Impact']
    return df

###########################
# 3. VISUALIZATION
###########################
def create_risk_distribution_colored_table(df):
    all_ratings = ['Very Low', 'Low', 'Medium', 'High', 'Extreme']
    risk_distribution = pd.crosstab(df['Type of Risk'], df['Rating'])
    risk_distribution = risk_distribution.reindex(columns=all_ratings, fill_value=0)
    risk_distribution = risk_distribution.replace(0, '-')
    row_sums = risk_distribution.replace('-', 0).sum(axis=1)
    risk_distribution['Total'] = row_sums

    rating_mapping = {'Very Low': 1, 'Low': 2, 'Medium': 3, 'High': 4, 'Extreme': 5}
    risk_colors = {level: RISK_COLORS[rating_mapping[level]][0] for level in all_ratings}

    font_colors = {'Very Low': 'black', 'Low': 'black', 'Medium': 'black', 'High': 'white', 'Extreme': 'white', 'Total': 'black'}

    fig = go.Figure(data=go.Table(
        header=dict(values=['Type of Risk'] + list(risk_distribution.columns), fill_color='lightgray', align='center', font=dict(color='black', size=14), line=dict(width=1, color='white'), height=50),
        cells=dict(values=[risk_distribution.index] + [risk_distribution[level] for level in risk_distribution.columns], fill_color=['lightgray'] + [[risk_colors[level]] * len(risk_distribution) for level in all_ratings] + ['lightgray'] * len(risk_distribution), align='center', font=dict(color=[font_colors[level] for level in risk_distribution.columns] + ['black'], size=14), line=dict(width=1, color='white'), height=40)
    ))

    fig.update_layout(height=600, margin=dict(l=50, r=50, t=50, b=50))
    return fig, risk_distribution

def create_risk_distribution(df):
    risk_level_dist = df['Rating'].value_counts().sort_index()
    fig_risk_level = go.Figure(data=[go.Bar(x=risk_level_dist.index, y=risk_level_dist.values, marker_color=[RISK_COLORS[i][0] for i in range(1, 6)], text=risk_level_dist.values, textposition='auto')])
    fig_risk_level.update_layout(title="Risk Level Distribution", xaxis_title="Risk Level", yaxis_title="Count", height=400)

    type_dist = df['Type of Risk'].value_counts()
    fig_type_dist = go.Figure(data=[go.Bar(x=type_dist.index, y=type_dist.values, marker_color='#1E3F66', text=type_dist.values, textposition='auto')])
    fig_type_dist.update_layout(title="Risk Distribution by Type", xaxis_title="Risk Type", yaxis_title="Count", height=400)

    return fig_risk_level, fig_type_dist

def create_trend_analysis(df):
    avg_scores = df.groupby('Type of Risk')['Risk_Score'].mean().sort_values(ascending=False)
    fig = go.Figure(data=[go.Bar(x=avg_scores.index, y=avg_scores.values, marker_color='#1E3F66', text=[f"{x:.2f}" for x in avg_scores.values], textposition='auto')])
    fig.update_layout(title="Average Risk Score", xaxis_title="Risk Type", yaxis_title="Average Risk Score", height=400)
    return fig

###########################
# 4. EXCEL REPORT GENERATION
###########################
def generate_excel_report(processed_df, risk_distribution, total_risks, extreme_risks, high_risks, avg_score):
    excel_buffer = BytesIO()
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        # Write processed data to Excel
        processed_df.to_excel(writer, sheet_name='Report', index=False)
        
        # Reset index and rename columns for Risk Distribution
        risk_distribution = risk_distribution.reset_index().rename(columns={'index': 'Type of Risk'})
        
        # Write risk distribution to Excel
        risk_distribution.to_excel(writer, sheet_name='Risk Distribution', index=False)
        
        # Create and write Risk Matrix to Excel
        risk_matrix = create_risk_matrix(processed_df)
        risk_matrix.to_excel(writer, sheet_name='Risk Matrix')
        
        # Apply formatting to Excel sheets
        workbook = writer.book
        format_risk_distribution_sheet(workbook['Risk Distribution'], risk_distribution.columns)
        format_report_sheet(workbook['Report'], processed_df)
        format_risk_matrix_sheet(workbook['Risk Matrix'])

    excel_buffer.seek(0)
    return excel_buffer

def create_risk_matrix(df):
    risk_matrix = pd.crosstab(df['Limitation'], df['Impact'], margins=True).fillna(0).astype(int)
    
    # Ensure all levels 1-5 are present
    all_levels = list(range(1, 6))
    for level in all_levels:
        if level not in risk_matrix.index[:-1]:
            risk_matrix.loc[level] = 0
        if level not in risk_matrix.columns[:-1]:
            risk_matrix[level] = 0
    
    # Sort the index and columns
    risk_matrix = risk_matrix.reindex(all_levels + ['Total'])
    risk_matrix = risk_matrix.reindex(columns=all_levels + ['Total'])
    
    # Replace zeros with dashes for the main matrix (excluding totals)
    risk_matrix_display = risk_matrix.copy()
    risk_matrix_display.iloc[:-1, :-1] = risk_matrix_display.iloc[:-1, :-1].replace(0, '-')
    
    return risk_matrix_display

def format_risk_distribution_sheet(sheet, columns):
    column_colors = {
        'Very Low': '006400',  # Deep green
        'Low': '90EE90',      # Light green
        'Medium': 'FFD700',    # Yellow
        'High': 'FF0000',     # Red
        'Extreme': '722F37'    # Wine
    }
    
    gray_color = 'D3D3D3'
    
    # Color headers with gray
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=1, column=col)
        cell.fill = PatternFill(start_color=gray_color, end_color=gray_color, fill_type='solid')
    
    # Map column names to their indices in the sheet
    column_indices = {col_name: idx + 1 for idx, col_name in enumerate(columns)}
    
    # Color each column based on its name
    for col_name, color in column_colors.items():
        if col_name in column_indices:
            col_idx = column_indices[col_name]
            for row in range(2, sheet.max_row + 1):  # Start from row 2 to skip header
                cell = sheet.cell(row=row, column=col_idx)
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                if col_name in ['High', 'Extreme']:
                    cell.font = cell.font.copy(color='FFFFFF')
    
    # Color Type of Risk and Total columns with gray
    for col_name in ['Type of Risk', 'Total']:
        if col_name in column_indices:
            col_idx = column_indices[col_name]
            for row in range(2, sheet.max_row + 1):
                cell = sheet.cell(row=row, column=col_idx)
                cell.fill = PatternFill(start_color=gray_color, end_color=gray_color, fill_type='solid')

def format_report_sheet(sheet, df):
    risk_color_mapping = {
        'Very Low': '006400',  # Deep green
        'Low': '90EE90',      # Light green
        'Medium': 'FFD700',   # Yellow
        'High': 'FF0000',     # Red
        'Extreme': '722F37'   # Wine
    }
    
    # Get the column index for 'Rating'
    rating_column = None
    for idx, col in enumerate(df.columns, 1):
        if col == 'Rating':
            rating_column = idx
            break
    
    if rating_column:
        for row_idx, rating in enumerate(df['Rating'], 2):  # Start from row 2 to skip header
            cell = sheet.cell(row=row_idx, column=rating_column)
            if rating in risk_color_mapping:
                cell.fill = PatternFill(start_color=risk_color_mapping[rating], end_color=risk_color_mapping[rating], fill_type='solid')
                if rating in ['High', 'Extreme']:
                    cell.font = cell.font.copy(color='FFFFFF')

def format_risk_matrix_sheet(sheet):
    def get_diagonal_color(row_idx, col_idx):
        if (row_idx, col_idx) in [(1,1), (2,1), (1,2)]:  # Deep green cells
            return '006400'  # Deep green
        elif (row_idx, col_idx) in [(3,1), (2,2), (1,3)]:  # Light green cells
            return '90EE90'  # Light green
        elif (row_idx, col_idx) in [(4,1), (3,2), (2,3), (1,4), (5,1), (4,2), (3,3), (2,4), (1,5)]:  # Yellow cells
            return 'FFD700'  # Yellow
        elif (row_idx, col_idx) in [(5,2), (4,3), (3,4), (2,5), (5,3), (4,4), (3,5)]:  # Red cells
            return 'FF0000'  # Red
        elif (row_idx, col_idx) in [(5,4), (4,5), (5,5)]:  # Wine cells
            return '722F37'  # Wine
        return None
    
    gray_color = 'D3D3D3'
    
    # Apply color coding to risk matrix cells
    for row in range(2, 7):  # 5x5 matrix + header row
        for col in range(2, 7):  # 5x5 matrix + header column
            cell = sheet.cell(row=row, column=col)
            color = get_diagonal_color(row-1, col-1)  # Adjust indices to 1-based
            if color:
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    
    # Apply gray background to headers and totals
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=1, column=col)
        cell.fill = PatternFill(start_color=gray_color, end_color=gray_color, fill_type='solid')
    
    for row in range(1, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=1)
        cell.fill = PatternFill(start_color=gray_color, end_color=gray_color, fill_type='solid')
    
    for row in range(1, 8):
        cell = sheet.cell(row=row, column=7)
        cell.fill = PatternFill(start_color=gray_color, end_color=gray_color, fill_type='solid')
    
    for col in range(1, 8):
        cell = sheet.cell(row=7, column=col)
        cell.fill = PatternFill(start_color=gray_color, end_color=gray_color, fill_type='solid')

###########################
# 5. MAIN DASHBOARD
###########################
st.markdown("""
    <div class='dashboard-header'>
        <h1>🎯 Risk Intelligence Dashboard </h1>
        <div class='dashboard-subtitle'>
            Comprehensive Risk Assessment & Analytics Platform
        </div>
    </div>
""", unsafe_allow_html=True)

uploaded_file = st.file_uploader("Upload Risk Assessment Data", type=['xlsx', 'csv'], help="Upload your risk assessment Excel file")

if uploaded_file is not None:
    try:
        df = pd.read_excel(uploaded_file) if uploaded_file.name.endswith('.xlsx') else pd.read_csv(uploaded_file)
        df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
        df = df.dropna(axis=1, how='all')
        df = df.dropna(axis=0, how='all')

        required_columns = ['Risk#', 'Type of Risk', 'Risk ID', 'Description', 'Lower Limit', 'Upper Limit', 'Response Level']
        missing_columns = [c for c in required_columns if c not in df.columns]
        if missing_columns:
            st.error(f"Missing required columns: {', '.join(missing_columns)}")
            st.stop()

        processed_df = process_data(df)
        processed_df = processed_df.drop(columns=['Risk_Color', 'Lower Limit', 'Upper Limit', 'Response Level'], errors='ignore')

        if processed_df is not None:
            col1, col2, col3, col4 = st.columns(4)
            total_risks = len(processed_df)
            extreme_risks = len(processed_df[processed_df['Rating'] == 'Extreme'])
            high_risks = len(processed_df[processed_df['Rating'] == 'High'])
            avg_score = processed_df['Risk_Score'].mean()

            col1.metric("Extreme Risks", extreme_risks, f"{(extreme_risks / total_risks * 100):.1f}%")
            col2.metric("High Risks", high_risks, f"{(high_risks / total_risks * 100):.1f}%")
            col3.metric("Avg Risk Score", f"{avg_score:.2f}")

            tab1, tab2 = st.tabs(["Overview", "Analysis"])
            
            with tab1:
                st.subheader("Report")
                st.dataframe(processed_df, use_container_width=True)
            
                st.subheader("Risk Distribution by Type")
                fig_colored_table, risk_distribution = create_risk_distribution_colored_table(processed_df)
                st.plotly_chart(fig_colored_table, use_container_width=True)

            with tab2:
                colA, colB = st.columns(2)
                fig_risk_level, fig_type_dist = create_risk_distribution(processed_df)
                colA.plotly_chart(fig_risk_level, use_container_width=True)
                colB.plotly_chart(fig_type_dist, use_container_width=True)
                st.plotly_chart(create_trend_analysis(processed_df), use_container_width=True)

            st.subheader("Report Generation")
            st.info("Download Raw Data Report")
            excel_buffer = generate_excel_report(processed_df, risk_distribution, total_risks, extreme_risks, high_risks, avg_score)
            st.download_button(
                label="📈 Download Excel Data",
                data=excel_buffer,
                file_name=f"risk_assessment_data_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
    except Exception as e:
        st.error(f"Error: {str(e)}")
        st.write("Please check your file format and try again.")
        st.write("Error details:", e)
else:
    st.info("Upload your risk assessment file to begin analysis")
    st.markdown("""
    ### Required Format:
    - Risk# 
    - Type of Risk 
    - Risk ID 
    - Description 
    - Lower Limit 
    - Upper Limit
    - Appetite score
    (Optional: Response Level, etc.)
    """)